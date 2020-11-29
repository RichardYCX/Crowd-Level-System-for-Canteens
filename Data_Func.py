# Done By Michael

from collections import OrderedDict
import openpyxl
import ast
from Compiled_Store_Func import *
from Prediction_Func import *
import Basic_Store_Func

from pprint import pprint

MenuFile = 'test_menu.xlsx'
HourFile = 'test_hours.xlsx'
WaitingFile = 'test_waiting_time.xlsx'
AveragePeopleFile = 'user_input.txt'

def Get_Menu(): #returns a dictionary in the format of [store_name][day][period][menu]
    overall_menu = OrderedDict()
    Periods = ['Breakfast', 'Lunch', 'Dinner']
    Days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    wb = openpyxl.load_workbook(filename=MenuFile)
    worksheets = wb.sheetnames
    for worksheet in worksheets:
        store = wb[worksheet]
        column=1
        week_menu=OrderedDict()
        for day in Days:
            day_menu=OrderedDict()
            for period in Periods:
                menu_items=[]
                for row in store.iter_rows(min_row=3, min_col=column, max_col=column):
                    for cell in row:
                        if cell.value != None:
                            menu_items.append(cell.value)
                day_menu[period]=menu_items
                column+=1
            week_menu[day]=day_menu
        overall_menu[worksheet]=week_menu
    return overall_menu

def Get_Hour(): #returns a dictionary in the format of [store_name][day][period][hours]
    overall_Hour = OrderedDict()
    Periods = ['Opening Hours','Breakfast', 'Lunch', 'Dinner']
    Days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    wb = openpyxl.load_workbook(filename=HourFile)
    worksheets = wb.sheetnames
    for worksheet in worksheets:
        store = wb[worksheet]
        column=1
        week_hour=OrderedDict()
        for day in Days:
            day_hour=OrderedDict()
            for period in Periods:
                hour_items=[]
                for row in store.iter_rows(min_row=3, min_col=column, max_col=column):
                    for cell in row:
                        if cell.value != None:
                            hour_items.append(cell.value)
                day_hour[period]=hour_items
                column+=1
            week_hour[day]=day_hour
        overall_Hour[worksheet]=week_hour
    return overall_Hour

def Get_Waiting_Time(): #returns a dictionary in the format of [store_name][day][waiting time]
    print("Function/Get_Waiting_Time")
    overall=OrderedDict()
    Days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    wb = openpyxl.load_workbook(filename=WaitingFile)
    worksheets=wb.sheetnames
    for worksheet in worksheets:
        store=wb[worksheet]
        column=1
        day_hour=OrderedDict()
        for day in Days:    
            day_time=[]
            for row in store.iter_rows(min_row=2, min_col=column, max_col=column):
                for cell in row:
                    if cell.value != None:
                        day_time.append(cell.value)
            day_hour[day]=day_time
            column+=1
        overall[worksheet]=day_hour
    return overall

def Get_Average_People_Data(): #returns a dictionary in the format of [store_name][day][time slot][number of people]
                               #time_slot eg. 0900, 1000
    file = open(AveragePeopleFile, 'r')
    contents = file.read()
    contents = ast.literal_eval(contents)
    file.close()
    return contents


def Write_Average_People_Data(store_name, number_of_people):  # rewrites the data in user_input.txt with new data
    day, time = Basic_Store_Func.findDay('now')
    Queue_data = Get_Average_People_Data()
    time = str(int(time.strftime("%H"))) + '00'
    if len(time) == 3:
        time = '0' + time
    Queue_data[store_name][day][time].append(number_of_people)
    with open('user_input.txt', 'w') as file:
        print(Queue_data, file=file)

